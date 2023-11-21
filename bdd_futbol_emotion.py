import mysql.connector
import openpyxl
from futbol_emotion import *

def conectar_bdd():
    config = {
        'user': 'root',
        'password': 'Antonio',
        'host': 'localhost',
        'database': 'futbol_emotion'
    }
    conexion = mysql.connector.connect(**config)
    return conexion
def cerrar_conexion(conexion):
    conexion.commit()
    conexion.close()
def insertar_datos():
    #Obtenemos datos del scraping
    lista_botas = extraer_informacion()
    #Obtenemos la conexion hacia la BDD
    conexion = conectar_bdd()
    #Crear cursor
    cursor=conexion.cursor()
    script = 'insert into botas (nombre,marca,imagen,url,descripcion,tipo_suela) values (%s,%s,%s,%s,%s,%s,%s)'
    for botas in lista_botas:
        cursor.execute(script,(botas['nombre'], botas['marca'], botas['imagen'], botas['url'], botas['descripcion'],botas['tipo_suela']))

    cerrar_conexion(conexion)
def extraer_datos():
     lista_botas = []
     conexion = conectar_bdd()
     cursor = conexion.cursor(dictionary=True)
     script = "select * from botas"
     columnas = cursor.column_names
     #Ejecutamos consulta de SQL
     cursor.execute(script)
     #Obtener datos de la consulta
     lista_botas = cursor.fetchall()
     cerrar_conexion(conexion)
     return lista_botas, columnas
def exportar_datos_excel():
     lista_botas,nombre_columnas = extraer_datos()
     #Creamos hoja excel
     documento = openpyxl.Workbook()
     hoja_1 = documento.active
     #Rellenar fila de nombre de campos
     for contador, nombre_columnas in enumerate(nombre_columnas, start=1):
         hoja_1.cell(row=1,column=contador,value=nombre_columnas)

     #Rellenar filas de los datos de cada campo
     for fila, botas in enumerate(lista_botas, start=2):
         for columnas,campos in enumerate(botas.values(), start=1):
             hoja_1.cell(row=fila, column=columnas, value=campos)

     documento.save("futbol_emotion.xlsx")
def eliminacion_registro_id(id):
    conexion = conectar_bdd()
    cursor = conexion.cursor()
    script = "DELETE FROM botas WHERE id = %s"
    id_eliminar = (id,)
    cursor.execute(script, id_eliminar)
    cerrar_conexion(conexion)
def insertar_registro_diccionario(diccionario):
    conexion = conectar_bdd()
    cursor = conexion.cursor()
    script = 'insert into botas (nombre,marca,imagen,url,descripcion,tipo_suela) values (%s,%s,%s,%s,%s,%s,%s)'
    cursor.execute(script,(diccionario['nombre'], diccionario['marca'], diccionario['imagen'], diccionario['url'],diccionario['descripcion'],diccionario['tipo_suela']))
    cerrar_conexion(conexion)


insertar_datos()
"""insertar_registro_diccionario({"nombre": "ULTRA PLAY IT NIÑO ", "marca": "PUMA ", "precio": "40.99 ", "imagen": "https://www.futbolemotion.com/imagesarticulos/199997/750/zapatilla-puma-ultra-play-it-nino-amarillo-fluor-0.webp", "url": "https://www.futbolemotion.com/es/comprar/zapatilla/puma/ultra-play-it-nino-yellow-blaze-white-black", "descripcion": "Hora de abrirse paso, hora de dejar todas las preocupaciones y mirar hacia el futuro, un nuevo año se acerca. Ha llegado el nuevo pack de botas de fútbol de Puma, con alegría y color. Con él, nuevos objetivos y tramos de las competiciones, los cuales llegan a su punto álgido. Puma envuelve las Future y Ultra en nuevas fuerzas que cuentan con explosiones de colores para el Voltage Pack. ", "tipo_suela": "Bota de nivel iniciación desarrollada para usarlos en campos de cemento y parquet. Orientada a jugadores que busquen una bota ligera y con un ajuste personalizado."})"""






